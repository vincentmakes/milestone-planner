"""
Export router for project exports.

Provides endpoints for:
- Exporting projects to MS Project XML format
- Exporting projects to CSV format

Uses MPXJ library for XML export when available.
"""

import csv
import io
import json
import re
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.middleware.auth import require_superuser
from app.models.custom_column import CustomColumn, CustomColumnValue
from app.models.equipment import Equipment, EquipmentBlock
from app.models.note import Note
from app.models.project import Project, ProjectPhase, ProjectSubphase
from app.models.site import BankHoliday, CompanyEvent, Site
from app.models.skill import Skill, UserSkill
from app.models.tag import Tag
from app.models.user import User, UserSite
from app.models.vacation import Vacation
from app.utils import utcnow_naive

router = APIRouter(tags=["export"])


def calculate_duration_days(start_date: date, end_date: date) -> int:
    """Calculate working days between two dates (simple calculation)."""
    if not start_date or not end_date:
        return 1
    delta = (end_date - start_date).days
    return max(1, delta)


def format_date_for_export(d: date) -> str:
    """Format date as MM/DD/YYYY for MS Project compatibility."""
    if not d:
        return ""
    if isinstance(d, str):
        d = datetime.fromisoformat(d).date()
    return d.strftime("%m/%d/%Y")


async def get_project_with_hierarchy(db: AsyncSession, project_id: int) -> Project | None:
    """Load project with all phases and subphases."""
    result = await db.execute(
        select(Project).options(selectinload(Project.phases)).where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()

    if not project:
        return None

    # Load subphases for each phase
    for phase in project.phases:
        subphases_result = await db.execute(
            select(ProjectSubphase)
            .where(ProjectSubphase.parent_id == phase.id, ProjectSubphase.parent_type == "phase")
            .order_by(ProjectSubphase.sort_order)
        )
        phase.subphases = list(subphases_result.scalars().all())

        # Recursively load nested subphases
        await load_nested_subphases(db, phase.subphases)

    return project


async def load_nested_subphases(db: AsyncSession, subphases: list[ProjectSubphase]):
    """Recursively load children of subphases."""
    for subphase in subphases:
        result = await db.execute(
            select(ProjectSubphase)
            .where(
                ProjectSubphase.parent_id == subphase.id, ProjectSubphase.parent_type == "subphase"
            )
            .order_by(ProjectSubphase.sort_order)
        )
        subphase.children = list(result.scalars().all())
        if subphase.children:
            await load_nested_subphases(db, subphase.children)


def build_task_list(project: Project) -> list[dict[str, Any]]:
    """Build flat task list from project hierarchy for export."""
    tasks = []
    task_id = 1

    # Project summary task (level 1)
    project_notes = []
    if project.customer:
        project_notes.append(f"Customer: {project.customer}")
    if hasattr(project, "pm") and project.pm:
        project_notes.append(f"PM: {project.pm.name}")

    tasks.append(
        {
            "id": task_id,
            "name": project.name,
            "outline_level": 1,
            "start": project.start_date,
            "finish": project.end_date,
            "duration": calculate_duration_days(project.start_date, project.end_date),
            "predecessors": "",
            "percent_complete": 0,
            "milestone": False,
            "notes": "; ".join(project_notes) if project_notes else "",
        }
    )
    task_id += 1

    # Sort phases by sort_order
    sorted_phases = sorted(project.phases, key=lambda p: p.sort_order or 0)

    # Add phases (level 2)
    for phase in sorted_phases:
        phase_completion = (
            phase.completion if hasattr(phase, "completion") and phase.completion else 0
        )

        # Parse dependencies
        deps_str = ""
        if phase.dependencies:
            try:
                deps = (
                    json.loads(phase.dependencies)
                    if isinstance(phase.dependencies, str)
                    else phase.dependencies
                )
                if deps:
                    deps_str = ",".join([f"{d.get('id', '')}{d.get('type', 'FS')}" for d in deps])
            except Exception:
                pass

        tasks.append(
            {
                "id": task_id,
                "name": phase.type,
                "outline_level": 2,
                "start": phase.start_date,
                "finish": phase.end_date,
                "duration": calculate_duration_days(phase.start_date, phase.end_date),
                "predecessors": deps_str,
                "percent_complete": phase_completion,
                "milestone": bool(phase.is_milestone),
                "notes": "",
            }
        )
        task_id += 1

        # Add subphases recursively
        if hasattr(phase, "subphases") and phase.subphases:
            task_id = add_subphases_to_list(tasks, phase.subphases, 3, task_id)

    return tasks


def add_subphases_to_list(
    tasks: list[dict], subphases: list[ProjectSubphase], level: int, task_id: int
) -> int:
    """Recursively add subphases to task list."""
    sorted_subphases = sorted(subphases, key=lambda s: s.sort_order or 0)

    for subphase in sorted_subphases:
        completion = (
            subphase.completion if hasattr(subphase, "completion") and subphase.completion else 0
        )

        # Parse dependencies
        deps_str = ""
        if subphase.dependencies:
            try:
                deps = (
                    json.loads(subphase.dependencies)
                    if isinstance(subphase.dependencies, str)
                    else subphase.dependencies
                )
                if deps:
                    deps_str = ",".join([f"{d.get('id', '')}{d.get('type', 'FS')}" for d in deps])
            except Exception:
                pass

        tasks.append(
            {
                "id": task_id,
                "name": subphase.name,
                "outline_level": level,
                "start": subphase.start_date,
                "finish": subphase.end_date,
                "duration": calculate_duration_days(subphase.start_date, subphase.end_date),
                "predecessors": deps_str,
                "percent_complete": completion,
                "milestone": bool(subphase.is_milestone),
                "notes": "",
            }
        )
        task_id += 1

        # Add children recursively
        if hasattr(subphase, "children") and subphase.children:
            task_id = add_subphases_to_list(tasks, subphase.children, level + 1, task_id)

    return task_id


def generate_csv(tasks: list[dict[str, Any]], project_name: str) -> str:
    """Generate CSV content from task list."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow(
        [
            "ID",
            "Name",
            "Outline Level",
            "Start",
            "Finish",
            "Duration",
            "Predecessors",
            "% Complete",
            "Milestone",
            "Notes",
        ]
    )

    # Data rows
    for task in tasks:
        writer.writerow(
            [
                task["id"],
                task["name"],
                task["outline_level"],
                format_date_for_export(task["start"]),
                format_date_for_export(task["finish"]),
                f"{task['duration']}d",
                task["predecessors"],
                task["percent_complete"],
                "Yes" if task["milestone"] else "No",
                task["notes"],
            ]
        )

    return output.getvalue()


def generate_xml(tasks: list[dict[str, Any]], project_name: str, project: Project) -> str:
    """Generate MS Project XML content from task list."""
    # Basic MS Project XML structure
    xml_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<Project xmlns="http://schemas.microsoft.com/project">',
        f"  <Name>{escape_xml(project_name)}</Name>",
        f"  <Title>{escape_xml(project_name)}</Title>",
        f"  <StartDate>{format_xml_date(project.start_date)}</StartDate>",
        f"  <FinishDate>{format_xml_date(project.end_date)}</FinishDate>",
        "  <CalendarUID>1</CalendarUID>",
        "  <Calendars>",
        "    <Calendar>",
        "      <UID>1</UID>",
        "      <Name>Standard</Name>",
        "      <IsBaseCalendar>true</IsBaseCalendar>",
        "    </Calendar>",
        "  </Calendars>",
        "  <Tasks>",
    ]

    for task in tasks:
        xml_lines.extend(
            [
                "    <Task>",
                f"      <UID>{task['id']}</UID>",
                f"      <ID>{task['id']}</ID>",
                f"      <Name>{escape_xml(task['name'])}</Name>",
                f"      <OutlineLevel>{task['outline_level']}</OutlineLevel>",
                f"      <Start>{format_xml_date(task['start'])}</Start>",
                f"      <Finish>{format_xml_date(task['finish'])}</Finish>",
                f"      <Duration>PT{task['duration'] * 8}H0M0S</Duration>",
                f"      <PercentComplete>{task['percent_complete']}</PercentComplete>",
                f"      <Milestone>{'1' if task['milestone'] else '0'}</Milestone>",
                f"      <Notes>{escape_xml(task['notes'])}</Notes>",
                "    </Task>",
            ]
        )

    xml_lines.extend(
        [
            "  </Tasks>",
            "</Project>",
        ]
    )

    return "\n".join(xml_lines)


def escape_xml(text: str) -> str:
    """Escape special XML characters."""
    if not text:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def format_xml_date(d: date) -> str:
    """Format date for MS Project XML."""
    if not d:
        return ""
    if isinstance(d, str):
        d = datetime.fromisoformat(d).date()
    return f"{d.isoformat()}T08:00:00"


@router.post("/export/mpp/{project_id}")
async def export_project_to_mpp(
    project_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_superuser),
):
    """
    Export a project to MS Project XML format.

    Returns an XML file that can be imported into Microsoft Project.

    Matches: POST /api/export/mpp/:projectId
    """
    project = await get_project_with_hierarchy(db, project_id)

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Build task list
    tasks = build_task_list(project)

    # Generate XML
    xml_content = generate_xml(tasks, project.name, project)

    # Return as file download
    filename = f"{project.name.replace(' ', '_')}.xml"

    return Response(
        content=xml_content,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export/csv/{project_id}")
async def export_project_to_csv(
    project_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_superuser),
):
    """
    Export a project to CSV format.

    Returns a CSV file compatible with MS Project import.

    Matches: POST /api/export/csv/:projectId
    """
    project = await get_project_with_hierarchy(db, project_id)

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Build task list
    tasks = build_task_list(project)

    # Generate CSV
    csv_content = generate_csv(tasks, project.name)

    # Return as file download
    filename = f"{project.name.replace(' ', '_')}.csv"

    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/csv/{project_id}")
async def export_project_to_csv_get(
    project_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_superuser),
):
    """
    Export a project to CSV format (GET version for direct download).

    Returns a CSV file compatible with MS Project import.
    """
    return await export_project_to_csv(project_id, db, user)


# =============================================================================
# Site data export (Excel)
# =============================================================================


def _safe_filename(name: str) -> str:
    """Sanitize a string for use in a filename."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
    return cleaned or "site"


def _iso(value: date | datetime | None) -> str:
    """Format a date/datetime for Excel cells."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return value.isoformat()


def _yes_no(value: int | bool | None) -> str:
    return "Yes" if value else "No"


async def build_site_export_workbook(db: AsyncSession, site_id: int) -> tuple[bytes, str]:
    """
    Build an .xlsx workbook with all data related to a site.

    Returns (file bytes, filename).
    """
    # Imported lazily so the rest of the module loads even if openpyxl isn't installed.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    site = (await db.execute(select(Site).where(Site.id == site_id))).scalar_one_or_none()
    if site is None:
        raise HTTPException(status_code=404, detail="Site not found")

    # --- Load all site-scoped data ---------------------------------------------------
    # Users linked to this site (via user_sites association)
    user_rows = (
        (
            await db.execute(
                select(User)
                .join(UserSite, UserSite.user_id == User.id)
                .where(UserSite.site_id == site_id)
                .order_by(User.last_name, User.first_name)
            )
        )
        .scalars()
        .all()
    )
    user_ids = {u.id for u in user_rows}

    # Projects on this site, with phases / subphases / assignments preloaded
    project_rows = (
        (
            await db.execute(
                select(Project)
                .where(Project.site_id == site_id)
                .options(
                    selectinload(Project.phases).selectinload(ProjectPhase.staff_assignments),
                    selectinload(Project.subphases).selectinload(ProjectSubphase.staff_assignments),
                    selectinload(Project.staff_assignments),
                    selectinload(Project.equipment_assignments),
                    selectinload(Project.tags),
                )
                .order_by(Project.name)
            )
        )
        .scalars()
        .all()
    )
    project_ids = [p.id for p in project_rows]

    # Equipment on this site
    equipment_rows = (
        (
            await db.execute(
                select(Equipment).where(Equipment.site_id == site_id).order_by(Equipment.name)
            )
        )
        .scalars()
        .all()
    )
    equipment_id_set = {e.id for e in equipment_rows}

    # Equipment blocks (maintenance / defect periods) for this site's equipment
    equipment_block_rows: list[EquipmentBlock] = []
    if equipment_id_set:
        equipment_block_rows = list(
            (
                await db.execute(
                    select(EquipmentBlock)
                    .where(EquipmentBlock.equipment_id.in_(equipment_id_set))
                    .order_by(EquipmentBlock.start_date)
                )
            )
            .scalars()
            .all()
        )

    # Skills are global - include all of them, plus per-user proficiency for the site's users
    skill_rows = (await db.execute(select(Skill).order_by(Skill.name))).scalars().all()
    user_skill_rows: list[UserSkill] = []
    if user_ids:
        user_skill_rows = list(
            (await db.execute(select(UserSkill).where(UserSkill.user_id.in_(user_ids))))
            .scalars()
            .all()
        )
    skill_by_id = {s.id: s for s in skill_rows}
    user_by_id = {u.id: u for u in user_rows}

    # Tags are global - include all of them; per-project assignments come from Project.tags
    tag_rows = (await db.execute(select(Tag).order_by(Tag.name))).scalars().all()

    # Vacations for users belonging to this site
    vacation_rows: list[Vacation] = []
    if user_ids:
        vacation_rows = list(
            (
                await db.execute(
                    select(Vacation)
                    .where(Vacation.staff_id.in_(user_ids))
                    .order_by(Vacation.start_date)
                )
            )
            .scalars()
            .all()
        )

    # Custom columns: site-specific + global, plus values for the projects/phases/subphases
    column_rows = (
        (
            await db.execute(
                select(CustomColumn)
                .where((CustomColumn.site_id == site_id) | (CustomColumn.site_id.is_(None)))
                .order_by(CustomColumn.display_order, CustomColumn.name)
            )
        )
        .scalars()
        .all()
    )
    column_ids = [c.id for c in column_rows]

    project_id_set = set(project_ids)
    phase_ids = [ph.id for p in project_rows for ph in p.phases]
    subphase_ids = [s.id for p in project_rows for s in p.subphases]

    column_values: list[CustomColumnValue] = []
    if column_ids:
        cv_query = select(CustomColumnValue).where(
            CustomColumnValue.custom_column_id.in_(column_ids)
        )
        column_values = list((await db.execute(cv_query)).scalars().all())
    # Filter values to entities belonging to this site
    relevant_values = [
        v
        for v in column_values
        if (v.entity_type == "project" and v.entity_id in project_id_set)
        or (v.entity_type == "phase" and v.entity_id in set(phase_ids))
        or (v.entity_type == "subphase" and v.entity_id in set(subphase_ids))
    ]
    column_by_id = {c.id: c for c in column_rows}

    # Bank holidays and company events for the site
    holiday_rows = (
        (
            await db.execute(
                select(BankHoliday).where(BankHoliday.site_id == site_id).order_by(BankHoliday.date)
            )
        )
        .scalars()
        .all()
    )
    event_rows = (
        (
            await db.execute(
                select(CompanyEvent)
                .where(CompanyEvent.site_id == site_id)
                .order_by(CompanyEvent.date)
            )
        )
        .scalars()
        .all()
    )
    note_rows = (
        (await db.execute(select(Note).where(Note.site_id == site_id).order_by(Note.date)))
        .scalars()
        .all()
    )

    # --- Build the workbook ----------------------------------------------------------
    wb = Workbook()
    # Drop the default sheet; we'll add named ones.
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    def add_sheet(title: str, headers: list[str], rows: list[list[Any]]) -> None:
        ws = wb.create_sheet(title=title[:31])  # Excel sheet name max 31 chars
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append(row)
        # Auto-size columns based on content (capped)
        for idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                if idx - 1 < len(row):
                    val = row[idx - 1]
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)
        ws.freeze_panes = "A2"

    # Site summary
    add_sheet(
        "Site",
        [
            "ID",
            "Name",
            "Location",
            "City",
            "Country code",
            "Region code",
            "Timezone",
            "Active",
            "Created at",
        ],
        [
            [
                site.id,
                site.name,
                site.location,
                site.city,
                site.country_code,
                site.region_code,
                site.timezone,
                _yes_no(site.active),
                _iso(site.created_at),
            ]
        ],
    )

    # Combined Projects / Phases / Subphases hierarchy.
    # Rows are emitted in tree order (project → phases → nested subphases) so
    # the spreadsheet reads top-to-bottom like the Gantt outline.
    subphases_by_parent: dict[tuple[str, int], list[ProjectSubphase]] = {}
    for p in project_rows:
        for sp in p.subphases:
            subphases_by_parent.setdefault((sp.parent_type, sp.parent_id), []).append(sp)
    for siblings in subphases_by_parent.values():
        siblings.sort(key=lambda s: (s.sort_order or 0, s.start_date))

    hierarchy_rows: list[list[Any]] = []

    def append_subphase_rows(project: Project, parent_key: tuple[str, int], depth: int) -> None:
        for sp in subphases_by_parent.get(parent_key, []):
            hierarchy_rows.append(
                [
                    "Subphase",
                    depth,
                    sp.id,
                    project.id,
                    project.name,
                    sp.parent_type,
                    sp.parent_id,
                    sp.name,
                    "",  # Customer
                    "",  # PM
                    "",  # Sales PM
                    "",  # Confirmed
                    "",  # Volume
                    _iso(sp.start_date),
                    _iso(sp.end_date),
                    sp.completion,
                    _yes_no(sp.is_milestone),
                    sp.sort_order,
                    sp.dependencies or "",
                    "",  # Archived
                    "",  # Notes
                    _iso(sp.created_at),
                ]
            )
            append_subphase_rows(project, ("subphase", sp.id), depth + 1)

    for p in project_rows:
        hierarchy_rows.append(
            [
                "Project",
                1,
                p.id,
                p.id,
                p.name,
                "",  # Parent type
                "",  # Parent ID
                p.name,
                p.customer,
                user_by_id[p.pm_id].full_name if p.pm_id and p.pm_id in user_by_id else "",
                p.sales_pm,
                _yes_no(p.confirmed),
                p.volume,
                _iso(p.start_date),
                _iso(p.end_date),
                "",  # Completion (project-level rollup not stored)
                "",  # Milestone
                "",  # Sort order
                "",  # Dependencies
                _yes_no(p.archived),
                p.notes,
                _iso(p.created_at),
            ]
        )
        for ph in sorted(p.phases, key=lambda x: (x.sort_order or 0, x.start_date)):
            hierarchy_rows.append(
                [
                    "Phase",
                    2,
                    ph.id,
                    p.id,
                    p.name,
                    "project",
                    p.id,
                    ph.type,
                    "",  # Customer
                    "",  # PM
                    "",  # Sales PM
                    "",  # Confirmed
                    "",  # Volume
                    _iso(ph.start_date),
                    _iso(ph.end_date),
                    ph.completion,
                    _yes_no(ph.is_milestone),
                    ph.sort_order,
                    ph.dependencies or "",
                    "",  # Archived
                    "",  # Notes
                    _iso(ph.created_at),
                ]
            )
            append_subphase_rows(p, ("phase", ph.id), 3)

    add_sheet(
        "Projects",
        [
            "Type",
            "Level",
            "ID",
            "Project ID",
            "Project name",
            "Parent type",
            "Parent ID",
            "Name",
            "Customer",
            "PM",
            "Sales PM",
            "Confirmed",
            "Volume",
            "Start date",
            "End date",
            "Completion %",
            "Milestone",
            "Sort order",
            "Dependencies",
            "Archived",
            "Notes",
            "Created at",
        ],
        hierarchy_rows,
    )

    # Users
    add_sheet(
        "Users",
        [
            "ID",
            "Email",
            "First name",
            "Last name",
            "Job title",
            "Role",
            "Max capacity %",
            "Active",
            "SSO provider",
            "Created at",
        ],
        [
            [
                u.id,
                u.email,
                u.first_name,
                u.last_name,
                u.job_title,
                u.role,
                u.max_capacity,
                _yes_no(u.active),
                u.sso_provider,
                _iso(u.created_at),
            ]
            for u in user_rows
        ],
    )

    # Equipment
    add_sheet(
        "Equipment",
        ["ID", "Name", "Type", "Description", "Active", "Created at"],
        [
            [e.id, e.name, e.type, e.description, _yes_no(e.active), _iso(e.created_at)]
            for e in equipment_rows
        ],
    )

    # Skills (global) and per-user proficiency
    add_sheet(
        "Skills",
        ["ID", "Name", "Description", "Color", "Created at"],
        [[s.id, s.name, s.description, s.color, _iso(s.created_at)] for s in skill_rows],
    )
    add_sheet(
        "User skills",
        ["User ID", "User name", "Skill ID", "Skill name", "Proficiency", "Assigned at"],
        [
            [
                us.user_id,
                user_by_id[us.user_id].full_name if us.user_id in user_by_id else "",
                us.skill_id,
                skill_by_id[us.skill_id].name if us.skill_id in skill_by_id else "",
                us.proficiency,
                _iso(us.assigned_at),
            ]
            for us in user_skill_rows
        ],
    )

    # Tags (global) and per-project tag assignments
    add_sheet(
        "Tags",
        ["ID", "Name", "Color", "Created at"],
        [[t.id, t.name, t.color, _iso(t.created_at)] for t in tag_rows],
    )
    project_tag_rows = [[p.id, p.name, t.id, t.name, t.color] for p in project_rows for t in p.tags]
    add_sheet(
        "Project tags",
        ["Project ID", "Project name", "Tag ID", "Tag name", "Tag color"],
        project_tag_rows,
    )

    # Vacations
    add_sheet(
        "Vacations",
        ["ID", "Staff ID", "Staff name", "Start date", "End date", "Description", "Created at"],
        [
            [
                v.id,
                v.staff_id,
                user_by_id[v.staff_id].full_name if v.staff_id in user_by_id else "",
                _iso(v.start_date),
                _iso(v.end_date),
                v.description,
                _iso(v.created_at),
            ]
            for v in vacation_rows
        ],
    )

    # Staff assignments (project / phase / subphase)
    project_assignment_data = []
    phase_assignment_data: list[list[Any]] = []
    subphase_assignment_data: list[list[Any]] = []
    project_by_id = {p.id: p for p in project_rows}
    phase_by_id = {ph.id: (p, ph) for p in project_rows for ph in p.phases}
    subphase_by_id = {sp.id: (p, sp) for p in project_rows for sp in p.subphases}

    for p in project_rows:
        for a in p.staff_assignments:
            project_assignment_data.append(
                [
                    a.id,
                    p.id,
                    p.name,
                    a.staff_id,
                    user_by_id[a.staff_id].full_name if a.staff_id in user_by_id else "",
                    a.allocation,
                    _iso(a.start_date),
                    _iso(a.end_date),
                    _iso(a.created_at),
                ]
            )
        for ph in p.phases:
            for pa in ph.staff_assignments:
                phase_assignment_data.append(
                    [
                        pa.id,
                        p.id,
                        p.name,
                        ph.id,
                        ph.type,
                        pa.staff_id,
                        user_by_id[pa.staff_id].full_name if pa.staff_id in user_by_id else "",
                        pa.allocation,
                    ]
                )
        for sp in p.subphases:
            for sa in sp.staff_assignments:
                subphase_assignment_data.append(
                    [
                        sa.id,
                        p.id,
                        p.name,
                        sp.id,
                        sp.name,
                        sa.staff_id,
                        user_by_id[sa.staff_id].full_name if sa.staff_id in user_by_id else "",
                        sa.allocation,
                    ]
                )

    add_sheet(
        "Project assignments",
        [
            "ID",
            "Project ID",
            "Project name",
            "Staff ID",
            "Staff name",
            "Allocation %",
            "Start date",
            "End date",
            "Created at",
        ],
        project_assignment_data,
    )
    add_sheet(
        "Phase assignments",
        [
            "ID",
            "Project ID",
            "Project name",
            "Phase ID",
            "Phase type",
            "Staff ID",
            "Staff name",
            "Allocation %",
        ],
        phase_assignment_data,
    )
    add_sheet(
        "Subphase assignments",
        [
            "ID",
            "Project ID",
            "Project name",
            "Subphase ID",
            "Subphase name",
            "Staff ID",
            "Staff name",
            "Allocation %",
        ],
        subphase_assignment_data,
    )

    # Equipment assignments
    equipment_by_id = {e.id: e for e in equipment_rows}
    equipment_assignment_data = []
    for p in project_rows:
        for ea in p.equipment_assignments:
            equipment_assignment_data.append(
                [
                    ea.id,
                    p.id,
                    p.name,
                    ea.equipment_id,
                    equipment_by_id[ea.equipment_id].name
                    if ea.equipment_id in equipment_by_id
                    else "",
                    _iso(ea.start_date),
                    _iso(ea.end_date),
                    _iso(ea.created_at),
                ]
            )
    add_sheet(
        "Equipment assignments",
        [
            "ID",
            "Project ID",
            "Project name",
            "Equipment ID",
            "Equipment name",
            "Start date",
            "End date",
            "Created at",
        ],
        equipment_assignment_data,
    )

    # Custom columns + values
    add_sheet(
        "Custom columns",
        [
            "ID",
            "Name",
            "Type",
            "List options",
            "Scope",
            "Display order",
            "Width",
            "Created at",
        ],
        [
            [
                c.id,
                c.name,
                c.column_type,
                c.list_options or "",
                "global" if c.site_id is None else f"site:{c.site_id}",
                c.display_order,
                c.width,
                _iso(c.created_at),
            ]
            for c in column_rows
        ],
    )

    custom_value_rows = []
    for v in relevant_values:
        col = column_by_id.get(v.custom_column_id)
        entity_name = ""
        if v.entity_type == "project":
            proj = project_by_id.get(v.entity_id)
            entity_name = proj.name if proj else ""
        elif v.entity_type == "phase":
            mapping = phase_by_id.get(v.entity_id)
            entity_name = mapping[1].type if mapping else ""
        elif v.entity_type == "subphase":
            mapping = subphase_by_id.get(v.entity_id)
            entity_name = mapping[1].name if mapping else ""
        custom_value_rows.append(
            [
                v.id,
                v.custom_column_id,
                col.name if col else "",
                v.entity_type,
                v.entity_id,
                entity_name,
                v.value or "",
                _iso(v.created_at),
            ]
        )
    add_sheet(
        "Custom column values",
        [
            "ID",
            "Column ID",
            "Column name",
            "Entity type",
            "Entity ID",
            "Entity name",
            "Value",
            "Created at",
        ],
        custom_value_rows,
    )

    # Bank holidays
    add_sheet(
        "Bank holidays",
        ["ID", "Date", "End date", "Name", "Year", "Custom", "Created at"],
        [
            [
                h.id,
                _iso(h.date),
                _iso(h.end_date),
                h.name,
                h.year,
                _yes_no(h.is_custom),
                _iso(h.created_at),
            ]
            for h in holiday_rows
        ],
    )

    # Company events
    add_sheet(
        "Company events",
        ["ID", "Date", "End date", "Name", "Color", "Created at"],
        [
            [e.id, _iso(e.date), _iso(e.end_date), e.name, e.color or "", _iso(e.created_at)]
            for e in event_rows
        ],
    )

    # Equipment blocks (maintenance / defect periods)
    equipment_name_by_id = {e.id: e.name for e in equipment_rows}
    add_sheet(
        "Equipment blocks",
        [
            "ID",
            "Equipment ID",
            "Equipment name",
            "Start date",
            "End date",
            "Reason",
            "Description",
            "Created at",
        ],
        [
            [
                b.id,
                b.equipment_id,
                equipment_name_by_id.get(b.equipment_id, ""),
                _iso(b.start_date),
                _iso(b.end_date),
                b.reason,
                b.description,
                _iso(b.created_at),
            ]
            for b in equipment_block_rows
        ],
    )

    # Staff notes
    add_sheet(
        "Staff notes",
        ["ID", "Date", "Text", "Type", "Staff ID", "Staff name", "Created at"],
        [
            [
                n.id,
                _iso(n.date),
                n.text,
                n.type,
                n.staff_id or "",
                user_by_id[n.staff_id].full_name if n.staff_id in user_by_id else "",
                _iso(n.created_at),
            ]
            for n in note_rows
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    today = utcnow_naive().strftime("%Y-%m-%d")
    filename = f"site_{_safe_filename(site.name)}_{today}.xlsx"
    return buffer.getvalue(), filename


@router.get("/export/site/{site_id}/excel")
async def export_site_to_excel(
    site_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_superuser),
):
    """
    Export all data for a site (projects, phases, users, equipment, skills,
    vacations, assignments, custom columns, holidays, events) as an Excel file.

    Restricted to admin / superuser. Superusers may only export sites they
    have access to.
    """
    if user.is_superuser and not user.is_admin and site_id not in user.site_ids:
        raise HTTPException(status_code=403, detail="Access denied to this site")

    content, filename = await build_site_export_workbook(db, site_id)
    return Response(
        content=content,
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
